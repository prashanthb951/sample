from django.shortcuts import render
import re
import PyPDF2 as p
import pdfplumber


from django.shortcuts import render
import openpyxl


def searchkey(request):
    path = "D:/ExcelProject/DQR-11-E-MAG05-EB05-ES-001-en-B-EQUIPMENT SPECIFICATION  Air Cooled Condenser (EB05)_GRA (" \
           "2).pdf "
    string = input("Enter Some Data:- ").upper()
    with pdfplumber.open(path) as pdf:
        page3 = pdf.pages[3]
        text = page3.extract_text()
    if re.findall(string,text):
        print(string)

    # print(text)
    new_vend_re = re.compile(r'^\S[A-Z].*')
    pdf_key = {}
    for line in text.split('\n'):
        if new_vend_re.match(line):
            a, *b = line.split()
            b = ' '.join(b)
    return render(request, 'index.html', {})



def index(request):
    if "GET" == request.method:
        return render(request, 'index.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)

        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)
        excel_data = list()
        for r in range(1,active_sheet.max_row+1):
            row_data = list()
            for c in range(1,active_sheet.max_column+1):
                print(active_sheet.cell(row=r, column=c).value)
                row_data.append(active_sheet.cell(r,c).value)
            excel_data.append(row_data)
        return render(request, 'index.html', {"excel_data":excel_data})